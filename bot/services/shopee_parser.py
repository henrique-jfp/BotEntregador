"""
🛵💨 SHOPEE ROMANEIO PARSER
Parser otimizado para romaneios reais da Shopee com lat/long embarcado
"""
import openpyxl
from typing import List, Dict, Optional
from dataclasses import dataclass
from pathlib import Path


@dataclass
class ShopeeDelivery:
    """Entrega individual parseada do romaneio"""
    tracking: str           # SPX TN
    address: str           # Destination Address
    bairro: str
    city: str
    zipcode: str
    latitude: float
    longitude: float
    sequence: int          # Ordem sugerida pela Shopee
    stop: int             # Stop group (múltiplas entregas no mesmo ponto)
    at_id: str            # Route ID


class ShopeeRomaneioParser:
    """Parser de Excel da Shopee → estrutura otimizada"""
    
    @staticmethod
    def parse(excel_path: str) -> List[ShopeeDelivery]:
        """
        Extrai entregas do romaneio Excel da Shopee
        
        Vantagens:
        - Lat/long JÁ VÊM no Excel (zero chamadas de geocoding!)
        - Stop groups identificam múltiplas entregas no mesmo prédio
        - Sequence dá dica da ordem sugerida (mas podemos otimizar)
        """
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        deliveries = []
        
        for row in range(2, ws.max_row + 1):
            # Pula linhas vazias
            if not ws.cell(row, 4).value:  # SPX TN
                continue
                
            delivery = ShopeeDelivery(
                tracking=str(ws.cell(row, 4).value),
                address=str(ws.cell(row, 5).value or ""),
                bairro=str(ws.cell(row, 6).value or ""),
                city=str(ws.cell(row, 7).value or ""),
                zipcode=str(ws.cell(row, 8).value or ""),
                latitude=float(ws.cell(row, 9).value or 0),
                longitude=float(ws.cell(row, 10).value or 0),
                sequence=int(ws.cell(row, 2).value or 0),
                stop=int(ws.cell(row, 3).value or 0),
                at_id=str(ws.cell(row, 1).value or "")
            )
            
            deliveries.append(delivery)
        
        return deliveries
    
    @staticmethod
    def group_by_stop(deliveries: List[ShopeeDelivery]) -> Dict[int, List[ShopeeDelivery]]:
        """
        Agrupa entregas por Stop
        
        Útil para identificar prédios com múltiplas entregas
        (fazer todas de uma vez = eficiência máxima)
        """
        groups = {}
        for d in deliveries:
            if d.stop not in groups:
                groups[d.stop] = []
            groups[d.stop].append(d)
        return groups
    
    @staticmethod
    def to_csv_format(deliveries: List[ShopeeDelivery], output_path: str):
        """
        Converte para formato CSV compatível com bot atual
        (para retrocompatibilidade)
        """
        import csv
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['tracking', 'endereco', 'lat', 'lon', 'prioridade'])
            
            for d in deliveries:
                # Monta endereço completo
                full_address = f"{d.address}, {d.bairro}, {d.city} - {d.zipcode}"
                writer.writerow([
                    d.tracking,
                    full_address,
                    d.latitude,
                    d.longitude,
                    'normal'  # Default
                ])
    
    @staticmethod
    def analyze_route(deliveries: List[ShopeeDelivery]) -> Dict[str, any]:
        """
        Análise estatística do romaneio
        
        Retorna insights úteis para otimização
        """
        from collections import Counter
        
        stops = [d.stop for d in deliveries]
        stop_counts = Counter(stops)
        
        # Calcula dispersão geográfica
        lats = [d.latitude for d in deliveries]
        lons = [d.longitude for d in deliveries]
        
        return {
            'total_deliveries': len(deliveries),
            'unique_stops': len(stop_counts),
            'multi_delivery_stops': sum(1 for count in stop_counts.values() if count > 1),
            'max_deliveries_per_stop': max(stop_counts.values()),
            'avg_deliveries_per_stop': len(deliveries) / len(stop_counts),
            'route_id': deliveries[0].at_id if deliveries else None,
            'geographic_center': {
                'lat': sum(lats) / len(lats),
                'lon': sum(lons) / len(lons)
            },
            'cities': list(set(d.city for d in deliveries)),
            'bairros': list(set(d.bairro for d in deliveries))
        }


# CLI para testar
if __name__ == "__main__":
    import sys
    import json
    
    if len(sys.argv) < 2:
        print("Uso: python shopee_parser.py <arquivo.xlsx>")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    
    print("🛵 SHOPEE PARSER - ANÁLISE TURBO\n")
    
    # Parse
    deliveries = ShopeeRomaneioParser.parse(excel_path)
    print(f"✅ {len(deliveries)} entregas parseadas\n")
    
    # Análise
    analysis = ShopeeRomaneioParser.analyze_route(deliveries)
    print("📊 ESTATÍSTICAS:")
    print(f"  Total entregas: {analysis['total_deliveries']}")
    print(f"  Stops únicos: {analysis['unique_stops']}")
    print(f"  Stops com múltiplas entregas: {analysis['multi_delivery_stops']}")
    print(f"  Máx entregas/stop: {analysis['max_deliveries_per_stop']}")
    print(f"  Média entregas/stop: {analysis['avg_deliveries_per_stop']:.1f}")
    print(f"  Rota ID: {analysis['route_id']}")
    print(f"  Cidades: {', '.join(analysis['cities'])}")
    print(f"  Bairros: {len(analysis['bairros'])}")
    
    # Grupos por stop
    print("\n📍 STOPS COM MÚLTIPLAS ENTREGAS:")
    groups = ShopeeRomaneioParser.group_by_stop(deliveries)
    for stop, items in sorted(groups.items(), key=lambda x: len(x[1]), reverse=True):
        if len(items) > 1:
            print(f"  Stop {stop}: {len(items)} entregas - {items[0].address}")
    
    # Exporta CSV
    csv_path = excel_path.replace('.xlsx', '_parsed.csv')
    ShopeeRomaneioParser.to_csv_format(deliveries, csv_path)
    print(f"\n💾 CSV exportado: {csv_path}")
