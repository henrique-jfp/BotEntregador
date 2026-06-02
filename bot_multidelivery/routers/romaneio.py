"""
Router para Importação de Romaneios (PDF, CSV, Shopee, etc.)
Suporta múltiplas importações na mesma sessão
"""
import logging
import io
import os
import tempfile
from fastapi import APIRouter, UploadFile, File, HTTPException, Query, Form
from bot_multidelivery.parsers.pdf_parser import extract_addresses_from_pdf
from bot_multidelivery.parsers.csv_parser import parse_csv_addresses
from bot_multidelivery.parsers.shopee_parser import parse_shopee_excel
from bot_multidelivery.parsers.md_romaneio_parser import parse_md_romaneio_excel
from bot_multidelivery.session import session_manager, Romaneio, DailySession
from bot_multidelivery.persistence import data_store
from bot_multidelivery.models import DeliveryPoint
from datetime import datetime
import uuid
import re

router = APIRouter(prefix="/romaneio", tags=["Import"])
logger = logging.getLogger(__name__)

@router.post("/import")
async def import_romaneio(
    file: UploadFile = File(None),
    route_value: float = Query(None),
    manual_addresses: str = Form(None)
):
    """
    Importa romaneio E/OU adiciona endereços manualmente
    Pode ser chamado múltiplas vezes para a mesma sessão
    """
    if not file and not manual_addresses:
        raise HTTPException(status_code=400, detail="Arquivo ou endereços obrigatórios")
    
    addresses = []
    session_id = None
    session = None
    
    try:
        # ===== IMPORTAR DE ARQUIVO =====
        if file:
            filename = file.filename.lower()
            content = await file.read()
            
            if filename.endswith('.pdf'):
                try:
                    # Shopee "PDF" function actually expects bytes and handles temp file creation
                    addresses = parse_shopee_pdf(content)
                    logger.info(f"✅ Shopee PDF: {len(addresses)} endereços")
                except Exception as e:
                    logger.warning(f"Falha no parser Shopee (Tentativa PDF): {e}")
                    # Fallback para PDF real
                    addresses = extract_addresses_from_pdf(io.BytesIO(content))
                    logger.info(f"✅ PDF genérico: {len(addresses)} endereços")
            
            elif filename.endswith('.csv'):
                addresses = parse_csv_addresses(io.BytesIO(content))
                logger.info(f"✅ CSV: {len(addresses)} endereços")
            
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                import openpyxl
                from io import BytesIO
                buffer = BytesIO(content)
                try:
                    wb = openpyxl.load_workbook(buffer)
                    try:
                        ws = wb.active
                        headers = [str(ws.cell(1, col).value or '').strip().lower() for col in range(1, ws.max_column + 1)]
                    finally:
                        wb.close()
                    buffer.seek(0)
                    if any('spx tn' in h or 'destination address' in h for h in headers):
                        addresses = parse_shopee_excel(buffer)
                        logger.info(f"✅ Excel Shopee: {len(addresses)} endereços")
                    elif any('endereço completo' in h or 'endereco completo' in h for h in headers):
                        addresses = parse_md_romaneio_excel(buffer)
                        logger.info(f"✅ Excel Nova Transportadora: {len(addresses)} endereços")
                    else:
                        raise HTTPException(status_code=400, detail="Modelo de Excel não reconhecido. Envie Shopee ou Transportadora.")
                except Exception as e:
                    logger.error(f"Erro no parser Excel: {e}")
                    raise HTTPException(status_code=400, detail=f"Erro ao processar Excel: {str(e)}")

            else:
                raise HTTPException(
                    status_code=400, 
                    detail="Formato não suportado. Use PDF, CSV ou Excel."
                )
        
        # ===== ADICIONAR ENDEREÇOS MANUALMENTE =====
        if manual_addresses:
            # Aceita múltiplos formatos colados:
            # - Uma linha por endereço ("Rua X, 100")
            # - Duas colunas por linha ("Bairro<TAB>Endereço" ou "Endereço<TAB>Bairro")
            #   Algumas planilhas (novo cliente) têm a coluna 'Endereço' como primeira coluna
            manual_list = []

            def is_probably_address(s: str) -> bool:
                if not s:
                    return False
                low = s.lower()
                # números normalmente indicam endereço (número da rua / apt)
                if any(ch.isdigit() for ch in s):
                    return True
                # palavras comuns em endereços
                address_tokens = ['rua', 'r.', 'avenida', 'av.', 'av', 'travessa', 'alameda', 'praça', 'praca', 'rod', 'rodovia', 'estrada', 'apto', 'apt', 'bloco']
                if any(t in low for t in address_tokens):
                    return True
                return False

            for raw in manual_addresses.split('\n'):
                line = raw.strip()
                if not line:
                    continue

                bairro = ''
                address_part = ''
                lat = None
                lon = None
                pkg_id = None

                # Tenta detectar separador de colunas (tab, ponto-e-vírgula, pipe)
                if '\t' in line:
                    parts = [p.strip() for p in line.split('\t') if p.strip()]
                elif ';' in line and line.count(';') <= 4:
                    parts = [p.strip() for p in line.split(';') if p.strip()]
                elif '|' in line:
                    parts = [p.strip() for p in line.split('|') if p.strip()]
                else:
                    parts = None

                # Se detectamos cabeçalho como "endereco"/"endereço"/"bairro", pular
                if parts and len(parts) >= 2:
                    header_tokens = [p.lower() for p in parts[:2]]
                    if any('endereco' in h or 'endereço' in h or 'end' in h or 'bairro' in h for h in header_tokens):
                        # cabeçalho detectado -> pular linha
                        # (ex: "Endereço\tBairro\tNF")
                        continue

                if parts and len(parts) >= 2:
                    # Heurística: detecta qual coluna é endereço
                    first, second = parts[0], ' '.join(parts[1:])
                    # Detecta modelo com NF,Bairro,Endereço,Latitude,Longitude
                    if len(parts) >= 5:
                        # tente extrair lat/lon das últimas duas colunas
                        try:
                            maybe_lat = float(parts[-2].replace(',', '.'))
                            maybe_lon = float(parts[-1].replace(',', '.'))
                            # Considera como lat/lon válidos
                            lat = maybe_lat
                            lon = maybe_lon
                            # Monta endereco a partir das colunas do meio
                            # Ex: [NF, Bairro, Endereço Completo, Latitude, Longitude]
                            # Endereço pode ocupar mais de uma coluna, então juntamos da 3ª até a penúltima
                            if len(parts) >= 5:
                                # NF na primeira coluna
                                pkg_id = parts[0]
                                bairro = parts[1]
                                address_part = ' '.join(parts[2:-2]).strip()
                        except Exception:
                            pass
                    if is_probably_address(first) and not is_probably_address(second):
                        address_part = first
                        # tenta pegar bairro na segunda coluna se for curto
                        bairro = parts[1] if len(parts) >= 2 else ''
                    elif is_probably_address(second) and not is_probably_address(first):
                        address_part = second
                        bairro = parts[0]
                    else:
                        # Ambíguo: assume primeira coluna é endereço (novo cliente pode usar esse formato)
                        # Se a primeira coluna for curta (ex: apenas bairro), tenta inverter
                        if len(first) > len(second):
                            address_part = first
                            bairro = second
                        else:
                            address_part = second
                            bairro = first
                else:
                    # Fallback: usa a linha inteira como endereço
                    address_part = line

                # Normalizar: preferimos um dict para que o fluxo posterior
                # consiga usar o campo 'bairro' ao geocodificar
                entry = {'bairro': bairro or '', 'address': address_part}
                if lat is not None and lon is not None:
                    entry['lat'] = lat
                    entry['lon'] = lon
                if pkg_id:
                    entry['id'] = pkg_id
                manual_list.append(entry)

            addresses.extend(manual_list)
            logger.info(f"✅ {len(manual_list)} endereços manuais adicionados")
        
        if not addresses:
            raise HTTPException(
                status_code=400,
                detail="Nenhum endereço encontrado"
            )
        
        # ===== CRIAR OU PEGAR SESSÃO =====
        # Se é primeira importação, criar nova sessão
        # Se não, verificar se existe sessão ativa para continuar
        
        # Por enquanto: sempre criar nova ou procurar sessão ativa
        from bot_multidelivery.session import session_manager
        active_session = session_manager.get_active_session()
        
        if active_session and not active_session.is_finalized:
            session = active_session
            session_id = session.session_id
            logger.info(f"📝 Continuando sessão existente: {session_id}")
        else:
            # Criar nova - TRUNCAR session_name para máximo 50 caracteres (limitação do banco)
            # TODO: Remover limite após migration do banco de 50 para 200
            filename = file.filename if file else "Manual"
            session_name = f"Romaneio: {filename}"
            if len(session_name) > 50:
                session_name = session_name[:47] + "..."  # Truncar e adicionar elipsis
            
            session = DailySession(
                session_name=session_name,
                date=datetime.now().strftime("%Y-%m-%d"),
                route_value=route_value or 0.0
            )
            session_id = session.session_id
            logger.info(f"🆕 Nova sessão criada: {session_id}")
        
        # ===== CONVERTER ENDEREÇOS EM PONTOS DE ENTREGA =====
        delivery_points = []
        rom_id = str(uuid.uuid4())[:8]
        
        for idx, addr in enumerate(addresses):
            # Normalização de dados (suporta Dict ou String)
            addr_text = addr
            pkg_id = f"{session_id}_pkg_{idx}"
            prio = "normal"
            
            # Variáveis para coordenadas
            addr_lat = 0.0
            addr_lng = 0.0
            addr_bairro = ""
            
            if isinstance(addr, dict):
                addr_text = addr.get('address', '')
                # Tenta pegar ID do pacote se existir
                if 'id' in addr:
                    pkg_id = str(addr['id'])
                if 'priority' in addr:
                    prio = addr['priority']
                # Usa coordenadas do Excel quando disponíveis
                if addr.get('lat') and addr.get('lon'):
                    try:
                        addr_lat = float(addr['lat'])
                        addr_lng = float(addr['lon'])
                    except (ValueError, TypeError):
                        pass
                # Pega bairro para geocoding
                if addr.get('bairro'):
                    addr_bairro = addr['bairro']
            
            if not addr_text:
                continue

            # Heurística: tentar extrair um código de rastreio da linha (ex: GRDN1044014160)
            # Se encontrarmos algo que pareça um tracking (letras+digitos, >=8 caracteres),
            # usamos isso como package_id para permitir bipagem direta.
            if not pkg_id or str(pkg_id).startswith(f"{session_id}_pkg_"):
                try:
                    txt = addr_text.upper()
                    # padrão: prefixos comuns (GRD, GRDN, AWB...) seguido de dígitos/letras
                    m = re.search(r"\b(GRD[A-Z0-9-]{4,}|GRDN[0-9A-Z-]{4,}|AWB[0-9A-Z-]{4,}|[A-Z0-9]{8,20})\b", txt)
                    if m:
                        candidate = m.group(0).replace('-', '').strip()
                        # requer pelo menos uma letra e um dígito para reduzir falsos positivos
                        if re.search(r"[A-Z]", candidate) and re.search(r"\d", candidate) and len(candidate) >= 8:
                            pkg_id = candidate
                except Exception:
                    pass

            # Se tivermos bairro separado (do input colado), anexa ao texto
            # para melhorar a chance de geocoding correto sem alterar DeliveryPoint.
            if addr_bairro:
                # Evita duplicar caso o bairro já esteja no texto
                if addr_bairro.strip().lower() not in addr_text.strip().lower():
                    addr_text = f"{addr_text}, {addr_bairro}"

            # Adaptação para o Dataclass DeliveryPoint
            # Usa coordenadas do Excel se disponíveis, senão será geocodificado depois
            try:
                point = DeliveryPoint(
                    address=addr_text,
                    lat=addr_lat,
                    lng=addr_lng,
                    romaneio_id=rom_id,
                    package_id=pkg_id,
                    priority=prio,
                    bairro=addr_bairro
                )
            except TypeError:
                # Fallback para definição antiga se houver mismatch de campos
                point = DeliveryPoint(
                    address=addr_text,
                    lat=0.0,
                    lng=0.0,
                    package_id=pkg_id, # Assume que package_id existe
                    # romaneio_id pode estar faltando em versões antigas
                )
                # Monkey patch se necessário
                if not hasattr(point, 'romaneio_id'):
                    point.romaneio_id = rom_id
                # Garantir bairro no fallback
                if not hasattr(point, 'bairro'):
                    point.bairro = addr_bairro

            delivery_points.append(point)
        
        # ===== ADICIONAR À SESSÃO =====
        rom = Romaneio(
            id=rom_id,
            uploaded_at=datetime.now(),
            points=delivery_points,
            filename=file.filename if file else "Manual"  # Rastreia o nome do arquivo
        )
        session.romaneios.append(rom)
        session.current_step = "importing"
        session_manager.save_session(session, set_as_current=True)
        session_manager.set_current_session(session_id)  # Garantir que está como sessão ativa
        
        logger.info(f"✅ Sessão {session_id} agora tem {sum(len(r.points) for r in session.romaneios)} pacotes totais")
        logger.info(f"📌 Sessão {session_id} definida como sessão ativa atual")
        
        return {
            "status": "success",
            "session_id": session_id,
            "romaneio_id": rom.id,
            "total_addresses": len(addresses),
            "session_total_packages": sum(len(r.points) for r in session.romaneios),
            "route_value": session.route_value,
            "imported_romaneios": len(session.romaneios),
            "message": f"Importado com sucesso: {len(addresses)} endereços"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao importar: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Erro: {str(e)}")


@router.get("/session/{session_id}/summary")
async def get_romaneio_summary(session_id: str):
    """
    Retorna resumo de todos os romaneios importados na sessão
    """
    try:
        session = session_manager.get_session(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Sessão não encontrada")
        
        return {
            "session_id": session_id,
            "session_name": session.session_name,
            "route_value": session.route_value,
            "total_romaneios": len(session.romaneios),
            "total_packages": sum(len(r.points) for r in session.romaneios),
            "romaneios": [
                {
                    "id": r.id,
                    "filename": r.filename,
                    "uploaded_at": r.uploaded_at.isoformat(),
                    "package_count": len(r.points),
                    "sample": [p.address for p in r.points[:3]]
                }
                for r in session.romaneios
            ]
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/romaneio/{session_id}/{romaneio_id}")
async def remove_romaneio(session_id: str, romaneio_id: str):
    """
    Remove um romaneio específico da sessão
    """
    try:
        session = session_manager.get_session(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Sessão não encontrada")
        
        session.romaneios = [r for r in session.romaneios if r.id != romaneio_id]
        session_manager.save_session(session)
        
        logger.info(f"✅ Romaneio {romaneio_id} removido da sessão {session_id}")
        
        return {
            "status": "success",
            "message": "Romaneio removido",
            "total_packages_remaining": sum(len(r.points) for r in session.romaneios)
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/session/{session_id}")
async def delete_session(session_id: str, force: bool = False):
    """
    Exclui uma sessão completamente incluindo:
    - Todos os pacotes, rotas, ganhos e custos
    """
    try:
        session = session_manager.get_session(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Sessão não encontrada")
        
        # Remove da memória (pode levantar se protegido)
        session_manager.delete_session(session_id, force=force)
        
        logger.info(f"🗑️ Sessão {session_id} excluída completamente")
        
        return {
            "status": "success",
            "message": "Sessão excluída com sucesso"
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{session_id}/reoptimize")
async def reoptimize_session(session_id: str, num_deliverers: int = None):
    """
    Re-otimiza a divisão e rotas para a sessão.
    Se `num_deliverers` for informado, usa esse valor, senão usa `session.num_deliverers`.
    Retorna resumo das novas rotas geradas.
    """
    try:
        session = session_manager.get_session(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Sessão não encontrada")

        # Coleta todos os pontos dos romaneios
        all_points = []
        for rom in session.romaneios:
            all_points.extend(rom.points)

        if not all_points:
            return {"status": "empty", "message": "Nenhum ponto para otimizar"}

        # Decide número de entregadores
        k = num_deliverers if num_deliverers and num_deliverers > 0 else (session.num_deliverers or max(1, len(all_points) // 20))

        from bot_multidelivery.clustering import TerritoryDivider
        from bot_multidelivery.session import Route as SessionRoute
        import uuid

        # Para entregas a pé, o modo 'pedestrian' usa distância em linha reta (Haversine)
        # e garante o balanceamento 50/50.
        routing_mode = "pedestrian"
        divider = TerritoryDivider(session.base_lat or 0.0, session.base_lng or 0.0, mode=routing_mode)
        clusters = divider.divide_into_clusters(all_points, k=k)

        new_routes = []
        colors = ['#FF4444', '#44FF44', '#4444FF', '#FFD700', '#FF69B4']
        for i, cluster in enumerate(clusters):
            optimized = divider.optimize_cluster_route(cluster)
            route = SessionRoute(id=str(uuid.uuid4()), cluster=cluster, color=colors[i % len(colors)])
            route.optimized_order = optimized
            new_routes.append(route)

        # PRESERVAR ATRIBUIÇÕES: tenta mapear entregadores antigos para novos clusters
        old_assignments = [r.assigned_to_telegram_id for r in session.routes if r.assigned_to_telegram_id]
        if old_assignments:
            # calcula centroides dos novos clusters
            def centroid_of(route_obj):
                pts = [p for p in route_obj.optimized_order]
                if not pts:
                    return (0.0, 0.0)
                return (sum(p.lat for p in pts) / len(pts), sum(p.lng for p in pts) / len(pts))

            assigned = {}
            remaining_deliverers = old_assignments.copy()
            # match greedy: para cada deliverer tenta encontrar cluster mais próximo do seu última posição anterior
            # pega último centro conhecido baseado nas rotas antigas
            old_centroids = []
            for r in session.routes:
                pts = [p for p in r.optimized_order]
                if pts:
                    old_centroids.append(((sum(p.lat for p in pts) / len(pts), sum(p.lng for p in pts) / len(pts)), r.assigned_to_telegram_id))

            # para cada novo route, escolha deliverer cujo old_centroid estiver mais próximo
            used_deliverers = set()
            for nr in new_routes:
                nc = centroid_of(nr)
                best = None
                best_dist = float('inf')
                for oc, did in old_centroids:
                    if did in used_deliverers:
                        continue
                    # haversine quick
                    from bot_multidelivery.clustering import haversine_distance
                    d = haversine_distance(nc[0], nc[1], oc[0], oc[1])
                    if d < best_dist:
                        best_dist = d
                        best = did
                if best is not None:
                    nr.assigned_to_telegram_id = best
                    used_deliverers.add(best)

        session.routes = new_routes
        session.current_step = 'optimized'
        session.num_deliverers = k
        session_manager.save_session(session)

        return {
            "status": "success",
            "session_id": session_id,
            "num_deliverers": k,
            "routes_created": len(new_routes),
            "total_packages": sum(len(r.optimized_order) for r in new_routes)
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao reotimizar: {e}")
        raise HTTPException(status_code=500, detail=str(e))
