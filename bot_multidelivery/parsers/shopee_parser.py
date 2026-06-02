"""
Parser Excel da Shopee - Extrai entregas do romaneio oficial
"""
from typing import List, Dict
from dataclasses import dataclass
import openpyxl
import re


def clean_destination_address(raw_address: str) -> str:
    """
    Limpa endereço da Shopee extraindo APENAS:
    - Nome da rua (antes da primeira vírgula)
    - Número do prédio (após a primeira vírgula, até encontrar espaço/vírgula/parêntese)
    
    Exemplos:
        "Rua Mena Barreto, 151, Portaria" -> "Rua Mena Barreto, 151"
        "Rua Principado de Mônaco, 37, Apt 501(guarita tb pode deixar" -> "Rua Principado de Mônaco, 37"
        "Rua Real Grandeza, 278, 601" -> "Rua Real Grandeza, 278"
    """
    if not raw_address:
        return ""
    
    # Remove espaços extras
    address = raw_address.strip()
    
    # Divide pela primeira vírgula
    parts = address.split(',', 2)  # Limita a 3 partes
    
    if len(parts) < 2:
        # Se não tem vírgula, retorna o endereço como está
        return address
    
    # Parte 1: Nome da rua
    street_name = parts[0].strip()
    
    # [FIX] Remove lixo comum que aparece no nome da rua (antes da vírgula)
    # Ex: "Rua X apt 201" -> "Rua X"
    # Remove qualquer coisa que pareça "apt", "ap", "bloco", "loja" seguido de digitos ou no final
    street_name = re.sub(r'\s+(?:apt\.?|ap\.?|apto\.?|bloco|bl\.?|loja|lj\.?|casa|sl\.?|sala)\s*.*$', '', street_name, flags=re.IGNORECASE)

    # Parte 2: Número do prédio (extrai apenas dígitos do início)
    number_part = parts[1].strip()
    
    # Extrai apenas o número (remove tudo após espaços, parênteses, vírgulas)
    number_match = re.match(r'^(\d+[A-Za-z]?)', number_part)
    if number_match:
        building_number = number_match.group(1)
    else:
        # Se não encontrar número, usa a parte toda
        building_number = number_part.split()[0] if ' ' in number_part else number_part
    
    # Retorna apenas rua + número
    return f"{street_name}, {building_number}"


@dataclass
class ShopeeDelivery:
    """Entrega individual do romaneio Shopee"""
    tracking: str
    address: str
    bairro: str
    city: str
    latitude: float
    longitude: float
    stop: int
    customer_name: str = ""
    phone: str = ""


def parse_shopee_excel(file_obj) -> list:
    """
    Parse Excel da Shopee (formato DD-MM-YYYY Nome.xlsx)
    Extrai:
    - Tracking code
    - Endereço completo
    - Lat/Lon (embutidos na planilha)
    - STOP (agrupamento de mesmo prédio)
    Returns:
        Lista de dicts com: id, address, lat, lon, stop
    """
    import logging
    logger = logging.getLogger("bot_multidelivery.parsers.shopee_parser")
    try:
        wb = openpyxl.load_workbook(file_obj)
        try:
            ws = wb.active
            headers = {}
            header_row = None
            for row in range(1, min(5, ws.max_row + 1)):
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row, col).value
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower().strip()
                        if any(x in cell_lower for x in ['spx tn', 'spx_tn', 'tracking', 'código', 'at id', 'atid']):
                            headers['tracking'] = col
                            header_row = row
                        elif any(x in cell_lower for x in ['destination', 'endereço', 'endereco', 'address']):
                            headers['address'] = col
                            if header_row is None:
                                header_row = row
                        elif any(x in cell_lower for x in ['rua', 'street']):
                            headers['street'] = col
                            if header_row is None:
                                header_row = row
                        elif any(x in cell_lower for x in ['bairro', 'distrito', 'district', 'neighborhood', 'neighbour']):
                            headers['bairro'] = col
                            if header_row is None:
                                header_row = row
                        elif 'city' in cell_lower or 'cidade' in cell_lower:
                            headers['city'] = col
                        elif 'latitude' in cell_lower or (cell_lower == 'lat'):
                            headers['lat'] = col
                        elif 'longitude' in cell_lower or (cell_lower in ['lng', 'lon', 'long']):
                            headers['lon'] = col
                        elif any(x in cell_lower for x in ['stop', 'parada', 'sequence']):
                            headers['stop'] = col
                        elif any(x in cell_lower for x in ['nome', 'cliente', 'customer', 'name']):
                            headers['customer'] = col
                        elif any(x in cell_lower for x in ['telefone', 'phone', 'tel']):
                            headers['phone'] = col
            if not header_row:
                logger.warning("Shopee: Cabeçalhos não encontrados. Formato inválido.")
                raise ValueError("Cabeçalhos não encontrados. Formato inválido.")
            if 'tracking' not in headers and 'address' not in headers and 'street' not in headers:
                logger.warning("Shopee: Não encontrei coluna de Tracking (SPX TN), Endereço (Destination) ou Rua.")
                raise ValueError("Não encontrei coluna de Tracking (SPX TN), Endereço (Destination) ou Rua.")
            addresses = []
            stop_counter = 1
            for row in range(header_row + 1, ws.max_row + 1):
                tracking = None
                if 'tracking' in headers:
                    tracking_cell = ws.cell(row, headers['tracking'])
                    if tracking_cell.value:
                        tracking = str(tracking_cell.value).strip()
                address = ""
                if 'address' in headers:
                    addr_cell = ws.cell(row, headers['address'])
                    if addr_cell.value:
                        address = str(addr_cell.value).strip()
                if not address and 'street' in headers:
                    street_cell = ws.cell(row, headers['street'])
                    street_val = str(street_cell.value or '').strip()
                    bairro_val = ""
                    if 'bairro' in headers:
                        bairro_cell = ws.cell(row, headers['bairro'])
                        bairro_val = str(bairro_cell.value or '').strip()
                    if street_val:
                        address = street_val if not bairro_val else f"{street_val}, {bairro_val}"
                if not tracking and not address:
                    logger.info(f"Shopee: Linha {row} ignorada (sem tracking e sem endereço)")
                    continue
                if not tracking:
                    tracking = f"PKG{row:04d}"
                bairro = ""
                if 'bairro' in headers:
                    bairro_cell = ws.cell(row, headers['bairro'])
                    if bairro_cell.value:
                        bairro = str(bairro_cell.value).strip()
                city = ""
                if 'city' in headers:
                    city_cell = ws.cell(row, headers['city'])
                    if city_cell.value:
                        city = str(city_cell.value).strip()
                lat = None
                lon = None
                if 'lat' in headers:
                    lat_cell = ws.cell(row, headers['lat']).value
                    if lat_cell:
                        try:
                            lat_str = str(lat_cell).strip().replace(',', '.')
                            lat = float(lat_str)
                        except:
                            logger.info(f"Shopee: Linha {row} latitude inválida: {lat_cell}")
                if 'lon' in headers:
                    lon_cell = ws.cell(row, headers['lon']).value
                    if lon_cell:
                        try:
                            lon_str = str(lon_cell).strip().replace(',', '.')
                            lon = float(lon_str)
                        except:
                            logger.info(f"Shopee: Linha {row} longitude inválida: {lon_cell}")
                stop = stop_counter
                if 'stop' in headers:
                    stop_cell = ws.cell(row, headers['stop']).value
                    if stop_cell:
                        try:
                            stop = int(stop_cell)
                        except:
                            logger.info(f"Shopee: Linha {row} stop inválido: {stop_cell}")
                customer = ""
                if 'customer' in headers:
                    customer = str(ws.cell(row, headers['customer']).value or '').strip()
                phone = ""
                if 'phone' in headers:
                    phone = str(ws.cell(row, headers['phone']).value or '').strip()
                if address:
                    cleaned_address = clean_destination_address(address)
                    addresses.append({
                        'id': tracking,
                        'address': cleaned_address,
                        'raw_address': address,
                        'lat': lat,
                        'lon': lon,
                        'stop': stop,
                        'bairro': bairro,
                        'city': city,
                        'customer': customer,
                        'phone': phone,
                        'tracking': tracking
                    })
                    logger.info(f"Shopee: Linha {row} importada - id={tracking}, endereço={cleaned_address}")
                    stop_counter += 1
            logger.info(f"Shopee: Total de endereços importados: {len(addresses)}")
            return addresses
        finally:
            wb.close()
    except Exception as e:
        logger.error(f"Shopee: Erro ao parsear Excel: {str(e)}")
        raise Exception(f"Erro ao parsear Excel Shopee: {str(e)}")
