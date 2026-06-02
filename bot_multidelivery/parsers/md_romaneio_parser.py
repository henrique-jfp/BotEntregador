import openpyxl
from typing import List, Dict

def parse_md_romaneio_excel(file_obj) -> list:
    """
    Parser para o modelo de romaneio da Nova Transportadora (colunas: NF, Bairro, Endereço Completo, Latitude, Longitude)
    Retorna lista de dicts com os campos padronizados.
    """
    wb = openpyxl.load_workbook(file_obj)
    try:
        ws = wb.active
        headers = {}
        header_row = None
        # Identifica cabeçalhos
        for row in range(1, min(5, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower().strip()
                    if cell_lower in ['nf', 'nº', 'nota']:
                        headers['nf'] = col
                        header_row = row
                    elif 'bairro' in cell_lower:
                        headers['bairro'] = col
                        if header_row is None:
                            header_row = row
                    elif 'endereço completo' in cell_lower or 'endereco completo' in cell_lower:
                        headers['endereco'] = col
                        if header_row is None:
                            header_row = row
                    elif 'latitude' in cell_lower:
                        headers['lat'] = col
                    elif 'longitude' in cell_lower:
                        headers['lon'] = col
        if not header_row or 'endereco' not in headers:
            raise ValueError('Cabeçalhos obrigatórios não encontrados.')
        addresses = []
        for row in range(header_row + 1, ws.max_row + 1):
            endereco = ws.cell(row, headers['endereco']).value
            if not endereco:
                continue
            bairro = ws.cell(row, headers['bairro']).value if 'bairro' in headers else ''
            lat = ws.cell(row, headers['lat']).value if 'lat' in headers else None
            lon = ws.cell(row, headers['lon']).value if 'lon' in headers else None
            try:
                lat = float(str(lat).replace(',', '.')) if lat else None
            except:
                lat = None
            try:
                lon = float(str(lon).replace(',', '.')) if lon else None
            except:
                lon = None
            addresses.append({
                'id': ws.cell(row, headers['nf']).value if 'nf' in headers else f'PKG{row:04d}',
                'address': endereco,
                'bairro': bairro,
                'lat': lat,
                'lon': lon
            })
        return addresses
    finally:
        wb.close()
