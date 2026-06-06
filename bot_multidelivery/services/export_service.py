"""
📄 SERVIÇO DE EXPORTAÇÃO - Foco Logístico
Exporta métricas de desempenho e motivos de falha para Excel e PDF
"""
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """Serviço de exportação de relatórios logísticos"""
    
    def __init__(self, output_dir: str = "data/exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(
        self, 
        session_id: str,
        session_name: str,
        packages: List[Dict],
        failure_stats: Dict[str, int]
    ) -> str:
        """
        Exporta resumo da sessão para Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            logger.error("openpyxl não instalado")
            raise ImportError("openpyxl é necessário para exportar Excel")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório Logístico"
        
        # Estilos
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título
        ws['A1'] = f"RELATÓRIO DE ENTREGAS - {session_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"ID da Sessão: {session_id}"
        ws['A3'] = f"Data de Exportação: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Seção de Resumo de Falhas
        ws['A5'] = "RESUMO DE INSUCESSOS"
        ws['A5'].font = Font(bold=True)
        row = 6
        ws.cell(row=row, column=1, value="Motivo").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Qtd").font = Font(bold=True)
        
        for reason, count in failure_stats.items():
            row += 1
            ws.cell(row=row, column=1, value=reason)
            ws.cell(row=row, column=2, value=count)
            
        # Lista de Pacotes
        row += 2
        ws.cell(row=row, column=1, value="DETALHAMENTO POR PACOTE").font = Font(bold=True)
        row += 1
        headers = ['ID Pacote', 'Endereço', 'Status', 'Motivo Falha', 'Entregador', 'Data Entrega']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
        for pkg in packages:
            row += 1
            ws.cell(row=row, column=1, value=pkg.get('id', '---'))
            ws.cell(row=row, column=2, value=pkg.get('address', '---'))
            ws.cell(row=row, column=3, value=pkg.get('status', '---'))
            ws.cell(row=row, column=4, value=pkg.get('failure_reason', '---'))
            ws.cell(row=row, column=5, value=pkg.get('deliverer_name', '---'))
            ws.cell(row=row, column=6, value=pkg.get('delivered_at', '---'))
            
        filename = f"relatorio_{session_id}_{datetime.now().strftime('%Y%md_%H%M')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        return str(filepath)

# Instância global
export_service = ExportService()
